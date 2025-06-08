import os
import random
import sqlite3
import hashlib
import re
from contextlib import contextmanager
from typing import List, Optional, Tuple, Dict
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory, make_response
from flask_wtf import CSRFProtect
from flask_talisman import Talisman
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from apscheduler.schedulers.background import BackgroundScheduler
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import logging
from werkzeug.utils import secure_filename
from html import escape
import atexit

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# ==============================================================================
# SECURITY CONFIGURATION
# ==============================================================================

def check_security_config():
    """Check critical security configuration on startup"""
    secret_key = os.getenv('SECRET_KEY')
    if not secret_key:
        raise RuntimeError("CRITICAL: SECRET_KEY environment variable must be set!")
    if len(secret_key) < 32:
        raise RuntimeError("CRITICAL: SECRET_KEY must be at least 32 characters long!")
    logger.info("✅ Security configuration checks passed")

# SECURE APP CONFIGURATION
is_production = os.getenv('FLASK_ENV') == 'production'
app.config.update(
    SECRET_KEY=os.getenv('SECRET_KEY'),
    MAX_CONTENT_LENGTH=16 * 1024 * 1024,
    SESSION_COOKIE_SECURE=is_production,  # Only True in production
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax' if is_production else None,  # More permissive in development
    PERMANENT_SESSION_LIFETIME=timedelta(hours=2),
    WTF_CSRF_ENABLED=is_production,  # Disable CSRF in development for easier testing
    WTF_CSRF_TIME_LIMIT=3600,
)

# Initialize security extensions
csrf = CSRFProtect(app)
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"]
)
limiter.init_app(app)

# Security headers
is_production = os.getenv('FLASK_ENV') == 'production'
talisman = Talisman(
    app,
    force_https=is_production,  # Only force HTTPS in production
    strict_transport_security=is_production,  # HSTS only in production
    content_security_policy={
        'default-src': "'self'",
        'script-src': ["'self'", "'unsafe-inline'"],  # Allow inline scripts in development
        'style-src': ["'self'", "'unsafe-inline'"],  # Allow inline styles in development
        'img-src': ["'self'", 'data:'],
    } if is_production else None,  # Disable CSP in development
    session_cookie_secure=is_production
)

# File upload configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Initialize scheduler
scheduler = BackgroundScheduler()
scheduler.start()
atexit.register(lambda: scheduler.shutdown())

# ==============================================================================
# INPUT VALIDATION & SANITIZATION FUNCTIONS
# ==============================================================================

def validate_logsheet_number(number):
    """Validate logsheet number format"""
    if not number or not isinstance(number, str):
        return False
    return bool(re.match(r'^\d{1,10}$', number.strip()))

def validate_q_number(number):
    """Validate Q number format"""
    if not number or not isinstance(number, str):
        return False
    return bool(re.match(r'^\d{1,10}$', number.strip()))

def validate_pilot_pin(pin):
    """Validate pilot PIN format"""
    if not pin or not isinstance(pin, str):
        return False
    return bool(re.match(r'^\d{5}$', pin.strip()))

def sanitize_text_input(text, max_length=100):
    """Sanitize text input"""
    if not text:
        return ""
    text = str(text).strip()[:max_length]
    return escape(text)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_admin_user():
    """Create admin user if none exists"""
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM pilots WHERE is_admin = TRUE")
    if cursor.fetchone()[0] == 0:
        admin_pin = str(random.randint(10000, 99999))
        cursor.execute('''
            INSERT INTO pilots (name, pin, personal_email, base_rate, is_admin)
            VALUES (?, ?, ?, ?, ?)
        ''', ('Admin', admin_pin, 'admin@company.com', 0.0, True))
        conn.commit()
        print(f"\n=== FROSTY'S ADMIN USER CREATED ===")
        print(f"PIN: {admin_pin}")
        print(f"Save this PIN - it's your admin access!")
        print("=" * 37)
    conn.close()

class DatabaseManager:
    def __init__(self):
        os.makedirs('data', exist_ok=True)
        self.db_path = os.path.join('data', 'pilot_portal.db')
        self.init_database()
        self.run_migrations()
    
    def init_database(self):
        """Initialize database with all required tables and columns"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Pilots table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pilots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                pin TEXT NOT NULL UNIQUE,
                personal_email TEXT NOT NULL,
                work_email TEXT,
                base_rate REAL NOT NULL,
                is_admin BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Aircraft table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS aircraft (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                registration TEXT NOT NULL UNIQUE,
                model TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Pilot-Aircraft assignments
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pilot_aircraft (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pilot_id INTEGER,
                aircraft_id INTEGER,
                FOREIGN KEY (pilot_id) REFERENCES pilots (id) ON DELETE CASCADE,
                FOREIGN KEY (aircraft_id) REFERENCES aircraft (id) ON DELETE CASCADE,
                UNIQUE(pilot_id, aircraft_id)
            )
        ''')
        
        # Flight entries table with ALL required columns
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS flight_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pilot_id INTEGER,
                aircraft_id INTEGER,
                other_pilot_id INTEGER,
                logsheet_number TEXT NOT NULL,
                q_number TEXT NOT NULL,
                flight_date DATE NOT NULL,
                flight_type TEXT NOT NULL CHECK(flight_type IN ('604', '704')),
                airtime REAL NOT NULL,
                flight_time REAL NOT NULL,
                short_notice BOOLEAN DEFAULT FALSE,
                rate_applied REAL NOT NULL,
                amount_earned REAL NOT NULL,
                photo_filename TEXT NOT NULL,
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                email_sent_at TIMESTAMP,
                FOREIGN KEY (pilot_id) REFERENCES pilots (id),
                FOREIGN KEY (aircraft_id) REFERENCES aircraft (id),
                FOREIGN KEY (other_pilot_id) REFERENCES pilots (id),
                UNIQUE(logsheet_number, q_number)
            )
        ''')
        
        # Operations emails table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ops_emails (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT NOT NULL UNIQUE,
                is_active BOOLEAN DEFAULT TRUE,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def run_migrations(self):
        """Run database migrations to add missing columns"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # Check existing columns in pilots table
            cursor.execute("PRAGMA table_info(pilots)")
            pilot_columns = [column[1] for column in cursor.fetchall()]
            
            if 'work_email' not in pilot_columns:
                cursor.execute('ALTER TABLE pilots ADD COLUMN work_email TEXT')
                logger.info("Added work_email column to pilots table")
            
            # Check existing columns in flight_entries table
            cursor.execute("PRAGMA table_info(flight_entries)")
            flight_columns = [column[1] for column in cursor.fetchall()]
            
            if 'photo_filename' not in flight_columns:
                cursor.execute('ALTER TABLE flight_entries ADD COLUMN photo_filename TEXT')
                cursor.execute('UPDATE flight_entries SET photo_filename = "legacy_no_photo.jpg" WHERE photo_filename IS NULL')
                logger.info("Added photo_filename column to flight_entries table")
            
            if 'email_sent_at' not in flight_columns:
                cursor.execute('ALTER TABLE flight_entries ADD COLUMN email_sent_at TIMESTAMP')
                logger.info("Added email_sent_at column to flight_entries table")
            
            if 'submitted_at' not in flight_columns:
                cursor.execute('ALTER TABLE flight_entries ADD COLUMN submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
                cursor.execute('UPDATE flight_entries SET submitted_at = CURRENT_TIMESTAMP WHERE submitted_at IS NULL')
                logger.info("Added submitted_at column to flight_entries table")
            
            conn.commit()
            
        except Exception as e:
            logger.error(f"Migration error: {e}")
            conn.rollback()
        finally:
            conn.close()
    
    def get_connection(self):
        return sqlite3.connect(self.db_path)

db_manager = DatabaseManager()

class RateCalculator:
    @staticmethod
    def calculate_rate_and_amount(pilot_id, flight_date, airtime, flight_type, short_notice):
        """Calculate rate and amount based on flight type and cumulative hours"""
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        # Get pilot's base rate
        cursor.execute("SELECT base_rate FROM pilots WHERE id = ?", (pilot_id,))
        result = cursor.fetchone()
        if not result:
            conn.close()
            return 0.0, 0.0
        base_rate = result[0]
        
        # For 604 flights, rate is always $0
        if flight_type == '604':
            conn.close()
            return 0.0, 0.0
        
        # For 704 flights, calculate based on cumulative hours
        month_start = flight_date.replace(day=1)
        cursor.execute('''
            SELECT COALESCE(SUM(airtime), 0) 
            FROM flight_entries 
            WHERE pilot_id = ? AND flight_date >= ? AND flight_date <= ?
        ''', (pilot_id, month_start, flight_date))
        
        cumulative_hours = cursor.fetchone()[0]
        conn.close()
        
        # Determine rate based on cumulative hours and conditions
        if cumulative_hours < 30.0:
            if short_notice:
                rate = 120.0
            else:
                rate = base_rate
        else:
            if short_notice:
                rate = 240.0  # Short notice + overtime
            else:
                rate = 120.0  # Overtime rate
        
        amount = airtime * rate
        return rate, amount

class ExcelGenerator:
    def generate_monthly_excel(self, pilot_id, year, month):
        """Generate Excel file for a pilot's monthly flights"""
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        # Get pilot info
        cursor.execute("SELECT name FROM pilots WHERE id = ?", (pilot_id,))
        pilot_result = cursor.fetchone()
        if not pilot_result:
            conn.close()
            return None
        pilot_name = pilot_result[0]
        
        # Get flight entries for the month
        month_start = datetime(year, month, 1).date()
        if month == 12:
            month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        cursor.execute('''
            SELECT fe.flight_date, a.registration, fe.logsheet_number, fe.q_number,
                   fe.flight_type, fe.airtime, fe.flight_time, fe.short_notice, 
                   fe.rate_applied, fe.amount_earned, p2.name as other_pilot
            FROM flight_entries fe
            JOIN aircraft a ON fe.aircraft_id = a.id
            LEFT JOIN pilots p2 ON fe.other_pilot_id = p2.id
            WHERE fe.pilot_id = ? AND fe.flight_date >= ? AND fe.flight_date <= ?
            ORDER BY fe.flight_date, a.registration
        ''', (pilot_id, month_start, month_end))
        
        entries = cursor.fetchall()
        conn.close()
        
        if not entries:
            return None
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        
        # Generate filename
        month_name = month_start.strftime("%B %Y")
        aircraft_list = list(set([entry[1] for entry in entries]))
        aircraft_str = "_".join(aircraft_list).replace('-', '')
        
        ws.title = f"{month_name} {aircraft_str}"
        
        # Headers
        headers = ["Date", "Aircraft", "Logsheet#", "Q#", "Type", "Airtime", "Flight Time", 
                  "Other Pilot", "Short Notice", "Rate Applied", "Amount Earned"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        total_amount = 0
        for row, entry in enumerate(entries, 2):
            ws.cell(row=row, column=1, value=entry[0].strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=entry[1])
            ws.cell(row=row, column=3, value=entry[2])
            ws.cell(row=row, column=4, value=entry[3])
            ws.cell(row=row, column=5, value=entry[4])
            ws.cell(row=row, column=6, value=entry[5])
            ws.cell(row=row, column=7, value=entry[6])
            ws.cell(row=row, column=8, value=entry[10] or "Solo")
            ws.cell(row=row, column=9, value="Yes" if entry[7] else "No")
            ws.cell(row=row, column=10, value=f"${entry[8]:.2f}")
            ws.cell(row=row, column=11, value=f"${entry[9]:.2f}")
            total_amount += entry[9]
        
        # Total row
        total_row = len(entries) + 2
        ws.cell(row=total_row, column=10, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=11, value=f"${total_amount:.2f}").font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filename = f"{pilot_name}_{month_name.replace(' ', '_')}_{aircraft_str}.xlsx"
        filepath = os.path.join('monthly_reports', filename)
        os.makedirs('monthly_reports', exist_ok=True)
        wb.save(filepath)
        
        return filepath

class EmailSender:
    def __init__(self):
        self.gmail_user = os.getenv('GMAIL_USER')
        self.gmail_password = os.getenv('GMAIL_PASSWORD')
    
    def send_logsheet_to_ops(self, flight_entry_id):
        """Send individual logsheet to ops with photo attachment"""
        if not self.gmail_user or not self.gmail_password:
            logger.error("Gmail credentials not configured")
            return False
            
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        try:
            # Get active ops emails
            cursor.execute("SELECT email FROM ops_emails WHERE is_active = TRUE")
            ops_emails = [row[0] for row in cursor.fetchall()]
            
            if not ops_emails:
                logger.warning("No active ops emails configured")
                conn.close()
                return True
            
            # Get flight entry details
            cursor.execute('''
                SELECT fe.id, fe.pilot_id, fe.aircraft_id, fe.other_pilot_id,
                       fe.logsheet_number, fe.q_number, fe.flight_date, fe.flight_type,
                       fe.airtime, fe.flight_time, fe.short_notice, fe.photo_filename,
                       fe.submitted_at, p.name as pilot_name, a.registration, 
                       p2.name as other_pilot_name
                FROM flight_entries fe
                JOIN pilots p ON fe.pilot_id = p.id
                JOIN aircraft a ON fe.aircraft_id = a.id
                LEFT JOIN pilots p2 ON fe.other_pilot_id = p2.id
                WHERE fe.id = ?
            ''', (flight_entry_id,))
            
            entry = cursor.fetchone()
            if not entry:
                logger.error(f"Flight entry {flight_entry_id} not found")
                conn.close()
                return False
            
            # Check if photo file exists
            photo_filename = entry[11]
            attachment_path = os.path.join(UPLOAD_FOLDER, photo_filename)
            
            if not os.path.exists(attachment_path):
                logger.error(f"Photo file not found: {attachment_path}")
                attachment_path = None
            
            # Create subject line with aircraft registration/tail number
            subject = f"Logsheet {entry[4]} - {entry[14]} - {entry[6]}"
            
            # Create email body
            body = f"""Flight Operations - Logsheet Submission

Pilot: {entry[13]}
Aircraft: {entry[14]}
Date: {entry[6]}
Logsheet #: {entry[4]}
Q Number: {entry[5]}
Flight Type: {entry[7]}
Airtime: {entry[8]} hours
Flight Time: {entry[9]} hours
Short Notice: {'Yes' if entry[10] else 'No'}
Other Pilot: {entry[15] if entry[15] else 'Solo'}

Submitted: {entry[12]}

{f"Logsheet photo ({photo_filename}) is attached." if attachment_path else "NOTE: Logsheet photo file was not found."}

---
Flight Operations
Professional Flight Services
"""
            
            all_success = True
            for ops_email in ops_emails:
                success = self._send_email_with_attachment(
                    to_email=ops_email,
                    subject=subject,
                    body=body,
                    attachment_path=attachment_path
                )
                if not success:
                    all_success = False
            
            if all_success:
                cursor.execute(
                    "UPDATE flight_entries SET email_sent_at = ? WHERE id = ?",
                    (datetime.now(), flight_entry_id)
                )
                conn.commit()
            
            conn.close()
            return all_success
            
        except Exception as e:
            logger.error(f"Failed to send logsheet to ops: {e}")
            conn.close()
            return False

    def send_monthly_reports(self, year, month):
        """Send monthly Excel reports to pilots"""
        if not self.gmail_user or not self.gmail_password:
            logger.error("Gmail credentials not configured")
            return False
            
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        # Get all pilots who have flights this month
        month_start = datetime(year, month, 1).date()
        if month == 12:
            month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        cursor.execute('''
            SELECT DISTINCT p.id, p.name, p.personal_email, p.work_email
            FROM pilots p
            JOIN flight_entries fe ON p.id = fe.pilot_id
            WHERE fe.flight_date >= ? AND fe.flight_date <= ? AND p.is_admin = FALSE
        ''', (month_start, month_end))
        
        pilots_with_flights = cursor.fetchall()
        conn.close()
        
        excel_generator = ExcelGenerator()
        success_count = 0
        
        # Generate and send reports for each pilot
        for pilot_id, pilot_name, personal_email, work_email in pilots_with_flights:
            try:
                filepath = excel_generator.generate_monthly_excel(pilot_id, year, month)
                if filepath:
                    # Send to work email if available, otherwise personal email
                    send_to_email = work_email if work_email else personal_email
                    success = self._send_email_with_attachment(
                        to_email=send_to_email,
                        subject=f"Monthly Flight Report - {pilot_name} - {month_start.strftime('%B %Y')}",
                        body=f"Dear {pilot_name},\n\nPlease find attached your monthly flight report for {month_start.strftime('%B %Y')}.\n\nBest regards,\nOperations Team",
                        attachment_path=filepath
                    )
                    if success:
                        success_count += 1
                        logger.info(f"Sent monthly report to {pilot_name} at {send_to_email}")
                    else:
                        logger.error(f"Failed to send report to {pilot_name}")
            except Exception as e:
                logger.error(f"Error processing report for {pilot_name}: {e}")
        
        logger.info(f"Monthly reports: {success_count}/{len(pilots_with_flights)} sent successfully")
        return success_count > 0

    def _send_email_with_attachment(self, to_email, subject, body, attachment_path=None):
        """Actually send the email with attachment"""
        try:
            msg = MIMEMultipart()
            msg['From'] = self.gmail_user
            msg['To'] = to_email
            msg['Subject'] = subject
            
            # Add the email body
            msg.attach(MIMEText(body, 'plain'))
            
            # Handle file attachment
            if attachment_path and os.path.exists(attachment_path):
                try:
                    with open(attachment_path, "rb") as attachment_file:
                        file_data = attachment_file.read()
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(file_data)
                    
                    encoders.encode_base64(part)
                    filename = os.path.basename(attachment_path)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename="{filename}"'
                    )
                    msg.attach(part)
                    
                except Exception as e:
                    logger.error(f"Failed to attach file {attachment_path}: {e}")
            
            # Send the email
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.gmail_user, self.gmail_password)
            text = msg.as_string()
            server.sendmail(self.gmail_user, to_email, text)
            server.quit()
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to send email to {to_email}: {e}")
            return False

# Security middleware
@app.before_request
def security_headers():
    """Add security headers to all responses."""
    # Skip for static files
    if request.endpoint == 'static':
        return None
        
    # Set security headers on the response
    response = make_response()
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    
    # Check for authentication on non-public routes
    public_routes = ['pilot_login', 'static']
    if request.endpoint not in public_routes and 'pilot_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('pilot_login'))
        
    return None

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response

# ==============================================================================
# PILOT PORTAL ROUTES
# ==============================================================================

@app.route('/')
def index():
    """Redirect root to login page."""
    return redirect(url_for('pilot_login'))

@app.route('/login', methods=['GET', 'POST'])
def pilot_login():
    """Handle login page display and authentication."""
    if request.method == 'POST':
        pin = request.form.get('pin', '').strip()
        
        if not validate_pilot_pin(pin):
            flash('Please enter a valid 5-digit PIN', 'error')
            logger.warning(f"Invalid PIN format attempted from {request.remote_addr}")
            return render_template('pilot_login.html')
        
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, is_admin FROM pilots WHERE pin = ?", (pin,))
        pilot = cursor.fetchone()
        conn.close()
        
        if pilot:
            session.permanent = True
            session['pilot_id'] = pilot[0]
            session['pilot_name'] = sanitize_text_input(pilot[1])
            session['is_admin'] = pilot[2]
            
            logger.info(f"Successful login: {pilot[1]} (ID: {pilot[0]})")
            
            if pilot[2]:
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('pilot_dashboard'))
        else:
            flash('Invalid PIN', 'error')
            logger.warning(f"Failed login attempt from {request.remote_addr}")
    
    # Handle GET request or failed login
    if request.args.get('force_logout'):
        session.clear()
    
    if 'pilot_id' in session:
        if session.get('is_admin'):
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('pilot_dashboard'))
        
    return render_template('pilot_login.html')


@app.route('/pilot_dashboard')
def pilot_dashboard():
    if 'pilot_id' not in session:
        return redirect(url_for('pilot_login'))
    
    pilot_id = session['pilot_id']
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    # Get current month stats
    now = datetime.now()
    month_start = now.replace(day=1).date()
    
    # Get all flight entries for the current month for debugging
    print(f"\nQuerying for flights since: {month_start}")
    cursor.execute('''
        SELECT id, flight_date, airtime, flight_time
        FROM flight_entries 
        WHERE pilot_id = ? AND flight_date >= ?
        ORDER BY flight_date
    ''', (pilot_id, month_start))
    
    all_entries = cursor.fetchall()
    print("\nDebug - All flight entries for the month:")
    print("ID\tDate\t	AT\tFT")
    print("-" * 50)
    for entry in all_entries:
        print(f"{entry[0]}\t{entry[1]}\t{entry[2]}\t{entry[3]}")
    
    # Get all stats in a single query to ensure consistency
    print("\nRunning stats query with params:", {"pilot_id": pilot_id, "month_start": month_start})
    cursor.execute('''
        SELECT 
            COUNT(*) as total_flights,
            COALESCE(SUM(airtime), 0) as total_airtime,
            COALESCE(SUM(flight_time), 0) as total_flight_time,
            COALESCE(SUM(amount_earned), 0) as total_earned
        FROM flight_entries 
        WHERE pilot_id = ? AND flight_date >= ?
    ''', (pilot_id, month_start))
    
    result = cursor.fetchone()
    print(f"\nRaw query result: {result}")
    
    total_flights = result[0]
    total_airtime = float(result[1])  # Ensure float conversion
    total_flight_time = float(result[2])  # Ensure float conversion
    total_earned = result[3]
    
    print(f"After conversion - Airtime: {total_airtime}, Flight Time: {total_flight_time}")
    
    print("\nDebug - Calculated Totals:")
    print(f"Total Flights: {total_flights}")
    print(f"Total Airtime: {total_airtime}")
    print(f"Total Flight Time: {total_flight_time}")
    print(f"Total Earned: {total_earned}")
    
    # Create stats dictionary
    stats = {
        'total_flights': total_flights,
        'total_airtime': total_airtime,
        'total_flight_time': total_flight_time,
        'total_earned': total_earned
    }
    
    print("\nSending to template:")
    print(f"Total Flights: {total_flights}")
    print(f"Total Airtime: {total_airtime} (type: {type(total_airtime)})")
    print(f"Total Flight Time: {total_flight_time} (type: {type(total_flight_time)})")
    print(f"Total Earned: {total_earned}")
    
    # Debug output
    print(f"\nDashboard Stats for {pilot_id} (since {month_start}):")
    print(f"Total flights: {total_flights}")
    print(f"Total airtime: {total_airtime}")
    print(f"Total flight time: {total_flight_time}")
    print(f"Total earned: {total_earned}")
    
    # Get recent flights
    cursor.execute('''
        SELECT fe.flight_date, a.registration, fe.logsheet_number, fe.q_number,
               fe.flight_type, fe.airtime, fe.flight_time
        FROM flight_entries fe
        JOIN aircraft a ON fe.aircraft_id = a.id
        WHERE fe.pilot_id = ?
        ORDER BY fe.flight_date DESC
        LIMIT 10
    ''', (pilot_id,))
    
    recent_flights = cursor.fetchall()
    conn.close()
    
    return render_template('pilot_dashboard.html', 
                         stats=stats,
                         recent_flights=recent_flights,
                         pilot_name=session['pilot_name'])

@app.route('/submit_logsheet')
def submit_logsheet():
    if 'pilot_id' not in session:
        return redirect(url_for('pilot_login'))
    
    session.pop('logsheet_data', None)
    session['logsheet_data'] = {}
    
    return redirect(url_for('logsheet_step', step='aircraft'))

@app.route('/logsheet/<step>', methods=['GET', 'POST'])
def logsheet_step(step):
    if 'pilot_id' not in session:
        return redirect(url_for('pilot_login'))
    
    pilot_id = session['pilot_id']
    logsheet_data = session.get('logsheet_data', {})
    
    if request.method == 'POST':
        # Save current step data with validation
        if step == 'aircraft':
            logsheet_data['aircraft_id'] = request.form.get('aircraft_id')
        elif step == 'date':
            logsheet_data['flight_date'] = request.form.get('flight_date')
        elif step == 'logsheet':
            logsheet_number = request.form.get('logsheet_number', '').strip()
            if not validate_logsheet_number(logsheet_number):
                flash('Please enter a valid logsheet number (numbers only)', 'error')
                return redirect(url_for('logsheet_step', step=step))
            logsheet_data['logsheet_number'] = logsheet_number
        elif step == 'q_number':
            q_number = request.form.get('q_number', '').strip()
            if not validate_q_number(q_number):
                flash('Please enter a valid Q number (numbers only)', 'error')
                return redirect(url_for('logsheet_step', step=step))
            logsheet_data['q_number'] = q_number
        elif step == 'flight_type':
            logsheet_data['flight_type'] = request.form.get('flight_type')
        elif step == 'airtime':
            try:
                airtime = float(request.form.get('airtime'))
                if airtime < 0 or airtime > 24:
                    raise ValueError("Invalid airtime range")
                logsheet_data['airtime'] = airtime
            except (ValueError, TypeError):
                flash('Please enter a valid airtime between 0 and 24 hours', 'error')
                return redirect(url_for('logsheet_step', step=step))
        elif step == 'flight_time':
            try:
                flight_time = float(request.form.get('flight_time'))
                if flight_time < 0 or flight_time > 24:
                    raise ValueError("Invalid flight time range")
                logsheet_data['flight_time'] = flight_time
            except (ValueError, TypeError):
                flash('Please enter a valid flight time between 0 and 24 hours', 'error')
                return redirect(url_for('logsheet_step', step=step))
        elif step == 'short_notice':
            logsheet_data['short_notice'] = request.form.get('short_notice') == 'yes'
        elif step == 'other_pilot':
            logsheet_data['other_pilot_id'] = request.form.get('other_pilot_id') or None
        elif step == 'photo':
            if 'photo' in request.files:
                file = request.files['photo']
                if file and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{pilot_id}_{file.filename}")
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    file.save(filepath)
                    logsheet_data['photo_filename'] = filename
                else:
                    flash('Please upload a valid photo/scan of your logsheet (PNG, JPG, or PDF)', 'error')
                    return redirect(url_for('logsheet_step', step=step))
            else:
                flash('Photo/scan of logsheet is required', 'error')
                return redirect(url_for('logsheet_step', step=step))
        
        session['logsheet_data'] = logsheet_data
        
        # Determine next step
        next_step = get_next_step(step)
        if next_step:
            return redirect(url_for('logsheet_step', step=next_step))
        else:
            return redirect(url_for('logsheet_confirm'))
    
    # GET request - show the form for current step
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    if step == 'aircraft':
        cursor.execute('''
            SELECT a.id, a.registration
            FROM aircraft a
            JOIN pilot_aircraft pa ON a.id = pa.aircraft_id
            WHERE pa.pilot_id = ?
            ORDER BY a.registration
        ''', (pilot_id,))
        aircraft_list = cursor.fetchall()
        conn.close()
        
        if not aircraft_list:
            flash('No aircraft assigned to your account. Contact admin.', 'error')
            return redirect(url_for('pilot_dashboard'))
        
        if len(aircraft_list) == 1:
            logsheet_data['aircraft_id'] = aircraft_list[0][0]
            session['logsheet_data'] = logsheet_data
            return redirect(url_for('logsheet_step', step='date'))
        
        return render_template('logsheet/aircraft.html', aircraft_list=aircraft_list)
    
    elif step == 'date':
        conn.close()
        today = datetime.now().date()
        return render_template('logsheet/date.html', today=today)
    
    elif step == 'logsheet':
        conn.close()
        return render_template('logsheet/logsheet.html')
    
    elif step == 'q_number':
        conn.close()
        return render_template('logsheet/q_number.html')
    
    elif step == 'flight_type':
        conn.close()
        return render_template('logsheet/flight_type.html')
    
    elif step == 'airtime':
        conn.close()
        return render_template('logsheet/airtime.html')
    
    elif step == 'flight_time':
        conn.close()
        return render_template('logsheet/flight_time.html')
    
    elif step == 'short_notice':
        conn.close()
        return render_template('logsheet/short_notice.html')
    
    elif step == 'other_pilot':
        aircraft_id = logsheet_data.get('aircraft_id')
        cursor.execute('''
            SELECT p.id, p.name
            FROM pilots p
            JOIN pilot_aircraft pa ON p.id = pa.pilot_id
            WHERE pa.aircraft_id = ? AND p.id != ? AND p.is_admin = FALSE
            ORDER BY p.name
        ''', (aircraft_id, pilot_id))
        other_pilots = cursor.fetchall()
        conn.close()
        
        return render_template('logsheet/other_pilot.html', other_pilots=other_pilots)
    
    elif step == 'photo':
        conn.close()
        return render_template('logsheet/photo.html')

def get_next_step(current_step):
    steps = ['aircraft', 'date', 'logsheet', 'q_number', 'flight_type', 
             'airtime', 'flight_time', 'short_notice', 'other_pilot', 'photo']
    try:
        current_index = steps.index(current_step)
        if current_index < len(steps) - 1:
            return steps[current_index + 1]
    except ValueError:
        pass
    return None

@app.route('/logsheet/confirm')
def logsheet_confirm():
    if 'pilot_id' not in session or 'logsheet_data' not in session:
        return redirect(url_for('pilot_login'))
    
    logsheet_data = session['logsheet_data']
    pilot_id = session['pilot_id']
    
    # Validate required data
    required_fields = ['aircraft_id', 'flight_date', 'logsheet_number', 'q_number', 
                      'flight_type', 'airtime', 'flight_time', 'short_notice', 'photo_filename']
    for field in required_fields:
        if field not in logsheet_data:
            flash('Missing required information. Please start over.', 'error')
            return redirect(url_for('submit_logsheet'))
    
    # Get display data
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    # Get aircraft name
    cursor.execute("SELECT registration FROM aircraft WHERE id = ?", (logsheet_data['aircraft_id'],))
    aircraft_result = cursor.fetchone()
    if not aircraft_result:
        flash('Invalid aircraft selected', 'error')
        conn.close()
        return redirect(url_for('submit_logsheet'))
    aircraft_name = aircraft_result[0]
    
    # Get other pilot name
    other_pilot_name = None
    if logsheet_data.get('other_pilot_id'):
        cursor.execute("SELECT name FROM pilots WHERE id = ?", (logsheet_data['other_pilot_id'],))
        result = cursor.fetchone()
        if result:
            other_pilot_name = result[0]
    
    conn.close()
    
    # Calculate rate and amount
    flight_date = datetime.strptime(logsheet_data['flight_date'], '%Y-%m-%d').date()
    rate, amount = RateCalculator.calculate_rate_and_amount(
        pilot_id, flight_date, logsheet_data['airtime'], 
        logsheet_data['flight_type'], logsheet_data['short_notice']
    )
    
    confirmation_data = {
        'aircraft': aircraft_name,
        'flight_date': logsheet_data['flight_date'],
        'logsheet_number': logsheet_data['logsheet_number'],
        'q_number': logsheet_data['q_number'],
        'flight_type': logsheet_data['flight_type'],
        'airtime': logsheet_data['airtime'],
        'flight_time': logsheet_data['flight_time'],
        'short_notice': logsheet_data['short_notice'],
        'other_pilot': other_pilot_name,
        'photo_uploaded': bool(logsheet_data.get('photo_filename')),
        'rate': rate,
        'amount': amount
    }
    
    return render_template('confirm.html', data=confirmation_data)

@app.route('/logsheet/submit', methods=['POST'])
def logsheet_submit():
    if 'pilot_id' not in session or 'logsheet_data' not in session:
        return redirect(url_for('pilot_login'))
    
    logsheet_data = session['logsheet_data']
    pilot_id = session['pilot_id']
    
    # Calculate rate and amount
    flight_date = datetime.strptime(logsheet_data['flight_date'], '%Y-%m-%d').date()
    rate, amount = RateCalculator.calculate_rate_and_amount(
        pilot_id, flight_date, logsheet_data['airtime'], 
        logsheet_data['flight_type'], logsheet_data['short_notice']
    )
    
    # Save to database
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            INSERT INTO flight_entries 
            (pilot_id, aircraft_id, other_pilot_id, logsheet_number, q_number, 
             flight_date, flight_type, airtime, flight_time, short_notice, 
             rate_applied, amount_earned, photo_filename)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (pilot_id, logsheet_data['aircraft_id'], logsheet_data.get('other_pilot_id'),
              logsheet_data['logsheet_number'], logsheet_data['q_number'], flight_date,
              logsheet_data['flight_type'], logsheet_data['airtime'], 
              logsheet_data['flight_time'], logsheet_data['short_notice'], rate, amount,
              logsheet_data['photo_filename']))
        
        flight_entry_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        # Send email to ops
        email_sender = EmailSender()
        email_success = email_sender.send_logsheet_to_ops(flight_entry_id)
        
        # Clear logsheet data
        session.pop('logsheet_data', None)
        
        if email_success:
            flash('Logsheet submitted successfully and sent to Frosty\'s ops!', 'success')
        else:
            flash('Logsheet submitted successfully! (Email delivery may have failed)', 'warning')
        return redirect(url_for('pilot_dashboard'))
        
    except sqlite3.IntegrityError:
        conn.close()
        flash('Error: This logsheet number and Q number combination already exists!', 'error')
        return redirect(url_for('logsheet_confirm'))

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully', 'info')
    
    response = redirect(url_for('pilot_login'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response

# ==============================================================================
# PILOT EARNINGS AND FLIGHTS ROUTES
# ==============================================================================

@app.route('/pilot_earnings')
def pilot_earnings():
    if 'pilot_id' not in session:
        return redirect(url_for('pilot_login'))
    
    pilot_id = session['pilot_id']
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    # Get current month stats
    now = datetime.now()
    month_start = now.replace(day=1).date()
    
    cursor.execute('''
        SELECT COUNT(*), COALESCE(SUM(airtime), 0), COALESCE(SUM(amount_earned), 0)
        FROM flight_entries 
        WHERE pilot_id = ? AND flight_date >= ?
    ''', (pilot_id, month_start))
    
    month_stats = cursor.fetchone()
    
    # Get pilot's base rate
    cursor.execute("SELECT base_rate FROM pilots WHERE id = ?", (pilot_id,))
    pilot_base_rate = cursor.fetchone()[0]
    
    # Get recent paid flights
    cursor.execute('''
        SELECT fe.flight_date, a.registration, fe.logsheet_number, fe.q_number,
               fe.flight_type, fe.airtime, fe.amount_earned, fe.short_notice, fe.rate_applied
        FROM flight_entries fe
        JOIN aircraft a ON fe.aircraft_id = a.id
        WHERE fe.pilot_id = ? AND fe.flight_type = '704'
        ORDER BY fe.flight_date DESC
        LIMIT 15
    ''', (pilot_id,))
    
    recent_paid_flights = cursor.fetchall()
    conn.close()
    
    # Unpack month_stats into individual variables
    total_flights = month_stats[0] if month_stats else 0
    total_hours = month_stats[1] if month_stats else 0.0
    total_earnings = month_stats[2] if month_stats else 0.0
    
    return render_template('pilot_earnings.html', 
                         total_flights=total_flights,
                         total_hours=total_hours,
                         total_earnings=total_earnings,
                         pilot_base_rate=pilot_base_rate,
                         recent_paid_flights=recent_paid_flights,
                         pilot_name=session['pilot_name'])

@app.route('/pilot_flights')
def pilot_flights():
    if 'pilot_id' not in session:
        return redirect(url_for('pilot_login'))
    
    pilot_id = session['pilot_id']
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT fe.flight_date, a.registration, fe.logsheet_number, fe.q_number,
               fe.flight_type, fe.airtime, fe.flight_time, fe.short_notice,
               p2.name as other_pilot, fe.submitted_at,
               CASE WHEN fe.pilot_id = ? THEN 'Primary' ELSE 'Secondary' END as pilot_role
        FROM flight_entries fe
        JOIN aircraft a ON fe.aircraft_id = a.id
        LEFT JOIN pilots p2 ON 
            CASE 
                WHEN fe.pilot_id = ? THEN fe.other_pilot_id 
                ELSE fe.pilot_id 
            END = p2.id
        WHERE fe.pilot_id = ? OR fe.other_pilot_id = ?
        ORDER BY fe.flight_date DESC
    ''', (pilot_id, pilot_id, pilot_id, pilot_id))
    
    all_flights = cursor.fetchall()
    conn.close()
    
    return render_template('pilot_flights.html', 
                         flights=all_flights,
                         pilot_name=session['pilot_name'])

# ==============================================================================
# ADMIN ROUTES
# ==============================================================================

@app.route('/admin')
def admin_dashboard():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    # Get summary stats
    cursor.execute("SELECT COUNT(*) FROM pilots WHERE is_admin = FALSE")
    total_pilots = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM flight_entries")
    total_flights = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM aircraft")
    total_aircraft = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM ops_emails WHERE is_active = TRUE")
    active_ops_emails = cursor.fetchone()[0]
    
    # Get total earnings
    cursor.execute("SELECT COALESCE(SUM(amount_earned), 0) FROM flight_entries")
    total_earnings = cursor.fetchone()[0] or 0
    
    # Get recent flights
    cursor.execute('''
        SELECT fe.flight_date, p.name, a.registration, fe.logsheet_number,
               fe.q_number, fe.flight_type, fe.airtime, fe.amount_earned
        FROM flight_entries fe
        JOIN pilots p ON fe.pilot_id = p.id
        JOIN aircraft a ON fe.aircraft_id = a.id
        ORDER BY fe.submitted_at DESC
        LIMIT 10
    ''')
    
    recent_flights = cursor.fetchall()
    
    # Get all pilots for the monthly summary dropdown
    cursor.execute("SELECT id, name FROM pilots WHERE is_admin = FALSE ORDER BY name")
    all_pilots = cursor.fetchall()
    
    conn.close()
    
    return render_template('admin/dashboard.html',
                         total_pilots=total_pilots,
                         total_flights=total_flights,
                         total_aircraft=total_aircraft,
                         active_ops_emails=active_ops_emails,
                         recent_flights=recent_flights,
                         all_pilots=all_pilots,
                         total_earnings=total_earnings,
                         current_month=datetime.now().strftime('%Y-%m'))

@app.route('/admin/pilots')
def admin_pilots():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT p.id, p.name, p.pin, p.personal_email, p.work_email, p.base_rate,
               GROUP_CONCAT(a.registration, ', ') as aircraft
        FROM pilots p
        LEFT JOIN pilot_aircraft pa ON p.id = pa.pilot_id
        LEFT JOIN aircraft a ON pa.aircraft_id = a.id
        WHERE p.is_admin = FALSE
        GROUP BY p.id, p.name, p.pin, p.personal_email, p.work_email, p.base_rate
        ORDER BY p.name
    ''')
    pilots_list = cursor.fetchall()
    conn.close()
    
    return render_template('admin/pilots.html', pilots=pilots_list)

@app.route('/admin/add_pilot', methods=['GET', 'POST'])
def admin_add_pilot():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    if request.method == 'POST':
        name = sanitize_text_input(request.form.get('name', '').strip())
        personal_email = request.form.get('personal_email', '').strip()
        work_email = request.form.get('work_email', '').strip()
        base_rate_str = request.form.get('base_rate', '').strip()
        aircraft_ids = request.form.getlist('aircraft_ids')
        
        if not name or not personal_email or not base_rate_str:
            flash('Name, personal email, and base rate are required', 'error')
            return redirect(url_for('admin_add_pilot'))
        
        try:
            base_rate = float(base_rate_str)
            if base_rate < 0:
                raise ValueError("Rate cannot be negative")
        except ValueError:
            flash('Invalid base rate - must be a positive number', 'error')
            return redirect(url_for('admin_add_pilot'))
        
        # Generate unique 5-digit PIN
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        max_attempts = 100
        for attempt in range(max_attempts):
            pin = f"{random.randint(10000, 99999)}"
            cursor.execute("SELECT COUNT(*) FROM pilots WHERE pin = ?", (pin,))
            if cursor.fetchone()[0] == 0:
                break
        else:
            flash('Unable to generate unique PIN. Please try again.', 'error')
            conn.close()
            return redirect(url_for('admin_add_pilot'))
        
        try:
            cursor.execute('''
                INSERT INTO pilots (name, pin, personal_email, work_email, base_rate)
                VALUES (?, ?, ?, ?, ?)
            ''', (name, pin, personal_email, work_email or None, base_rate))
            
            pilot_id = cursor.lastrowid
            
            # Assign aircraft
            for aircraft_id in aircraft_ids:
                cursor.execute('''
                    INSERT INTO pilot_aircraft (pilot_id, aircraft_id)
                    VALUES (?, ?)
                ''', (pilot_id, aircraft_id))
            
            conn.commit()
            flash(f'Pilot {name} added successfully! PIN: {pin}', 'success')
            
        except sqlite3.IntegrityError as e:
            if 'name' in str(e):
                flash('Error: Pilot name already exists!', 'error')
            elif 'personal_email' in str(e):
                flash('Error: Email already exists!', 'error')
            else:
                flash('Error: Unable to add pilot!', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('admin_pilots'))
    
    # GET request - show form
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, registration, model FROM aircraft ORDER BY registration")
    aircraft_list = cursor.fetchall()
    conn.close()
    
    return render_template('admin/add_pilot.html', aircraft_list=aircraft_list)

@app.route('/admin/edit_pilot/<int:pilot_id>', methods=['GET', 'POST'])
def admin_edit_pilot(pilot_id):
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        name = sanitize_text_input(request.form.get('name', '').strip())
        personal_email = request.form.get('personal_email', '').strip()
        work_email = request.form.get('work_email', '').strip()
        base_rate_str = request.form.get('base_rate', '').strip()
        aircraft_ids = request.form.getlist('aircraft_ids')
        
        if not name or not personal_email or not base_rate_str:
            flash('Name, personal email, and base rate are required', 'error')
            return redirect(url_for('admin_edit_pilot', pilot_id=pilot_id))
        
        try:
            base_rate = float(base_rate_str)
            if base_rate < 0:
                raise ValueError("Rate cannot be negative")
        except ValueError:
            flash('Invalid base rate - must be a positive number', 'error')
            return redirect(url_for('admin_edit_pilot', pilot_id=pilot_id))
        
        try:
            cursor.execute('''
                UPDATE pilots 
                SET name = ?, personal_email = ?, work_email = ?, base_rate = ?
                WHERE id = ? AND is_admin = FALSE
            ''', (name, personal_email, work_email or None, base_rate, pilot_id))
            
            # Remove existing aircraft assignments
            cursor.execute('DELETE FROM pilot_aircraft WHERE pilot_id = ?', (pilot_id,))
            
            # Add new aircraft assignments
            for aircraft_id in aircraft_ids:
                cursor.execute('''
                    INSERT INTO pilot_aircraft (pilot_id, aircraft_id)
                    VALUES (?, ?)
                ''', (pilot_id, aircraft_id))
            
            conn.commit()
            flash(f'Pilot {name} updated successfully!', 'success')
            
        except sqlite3.IntegrityError:
            flash('Error: Pilot name or email already exists!', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('admin_pilots'))
    
    # GET request - show form with existing data
    cursor.execute('''
        SELECT name, pin, personal_email, work_email, base_rate
        FROM pilots WHERE id = ? AND is_admin = FALSE
    ''', (pilot_id,))
    pilot = cursor.fetchone()
    
    if not pilot:
        flash('Pilot not found', 'error')
        conn.close()
        return redirect(url_for('admin_pilots'))
    
    # Get all aircraft
    cursor.execute("SELECT id, registration, model FROM aircraft ORDER BY registration")
    aircraft_list = cursor.fetchall()
    
    # Get pilot's assigned aircraft
    cursor.execute('''
        SELECT aircraft_id FROM pilot_aircraft WHERE pilot_id = ?
    ''', (pilot_id,))
    assigned_aircraft = [row[0] for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('admin/edit_pilot.html', 
                         pilot=pilot, 
                         aircraft_list=aircraft_list,
                         assigned_aircraft=assigned_aircraft,
                         pilot_id=pilot_id)

@app.route('/admin/delete_pilot/<int:pilot_id>')
def admin_delete_pilot(pilot_id):
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT name FROM pilots WHERE id = ? AND is_admin = FALSE", (pilot_id,))
    pilot = cursor.fetchone()
    
    if pilot:
        cursor.execute("DELETE FROM pilots WHERE id = ? AND is_admin = FALSE", (pilot_id,))
        conn.commit()
        flash(f'Pilot {pilot[0]} deleted successfully!', 'success')
    else:
        flash('Pilot not found', 'error')
    
    conn.close()
    return redirect(url_for('admin_pilots'))

@app.route('/admin/aircraft')
def admin_aircraft():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT a.id, a.registration, a.model,
               GROUP_CONCAT(p.name, ', ') as pilots
        FROM aircraft a
        LEFT JOIN pilot_aircraft pa ON a.id = pa.aircraft_id
        LEFT JOIN pilots p ON pa.pilot_id = p.id AND p.is_admin = FALSE
        GROUP BY a.id, a.registration, a.model
        ORDER BY a.registration
    ''')
    aircraft_list = cursor.fetchall()
    conn.close()
    
    return render_template('admin/aircraft.html', aircraft=aircraft_list)

@app.route('/admin/add_aircraft', methods=['GET', 'POST'])
def admin_add_aircraft():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    if request.method == 'POST':
        registration = sanitize_text_input(request.form.get('registration', '').strip().upper())
        model = sanitize_text_input(request.form.get('model', '').strip())
        
        if not registration:
            flash('Aircraft registration is required', 'error')
            return redirect(url_for('admin_add_aircraft'))
        
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO aircraft (registration, model)
                VALUES (?, ?)
            ''', (registration, model or None))
            conn.commit()
            flash(f'Aircraft {registration} added successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: Aircraft registration already exists!', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('admin_aircraft'))
    
    return render_template('admin/add_aircraft.html')

@app.route('/admin/edit_aircraft/<int:aircraft_id>', methods=['GET', 'POST'])
def admin_edit_aircraft(aircraft_id):
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        registration = sanitize_text_input(request.form.get('registration', '').strip().upper())
        model = sanitize_text_input(request.form.get('model', '').strip())
        
        if not registration:
            flash('Aircraft registration is required', 'error')
            return redirect(url_for('admin_edit_aircraft', aircraft_id=aircraft_id))
        
        try:
            cursor.execute('''
                UPDATE aircraft 
                SET registration = ?, model = ?
                WHERE id = ?
            ''', (registration, model or None, aircraft_id))
            conn.commit()
            flash(f'Aircraft {registration} updated successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: Aircraft registration already exists!', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('admin_aircraft'))
    
    # GET request - show form with existing data
    cursor.execute('SELECT registration, model FROM aircraft WHERE id = ?', (aircraft_id,))
    aircraft = cursor.fetchone()
    
    if not aircraft:
        flash('Aircraft not found', 'error')
        conn.close()
        return redirect(url_for('admin_aircraft'))
    
    conn.close()
    return render_template('admin/edit_aircraft.html', aircraft=aircraft, aircraft_id=aircraft_id)

@app.route('/admin/delete_aircraft/<int:aircraft_id>')
def admin_delete_aircraft(aircraft_id):
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT registration FROM aircraft WHERE id = ?", (aircraft_id,))
    aircraft = cursor.fetchone()
    
    if aircraft:
        cursor.execute("DELETE FROM aircraft WHERE id = ?", (aircraft_id,))
        conn.commit()
        flash(f'Aircraft {aircraft[0]} deleted successfully!', 'success')
    else:
        flash('Aircraft not found', 'error')
    
    conn.close()
    return redirect(url_for('admin_aircraft'))

@app.route('/admin/flights')
def admin_flights():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT fe.id, fe.flight_date, p.name, a.registration, fe.logsheet_number,
               fe.q_number, fe.flight_type, fe.airtime, fe.flight_time,
               fe.short_notice, fe.rate_applied, fe.amount_earned,
               p2.name as other_pilot, fe.submitted_at, fe.email_sent_at
        FROM flight_entries fe
        JOIN pilots p ON fe.pilot_id = p.id
        JOIN aircraft a ON fe.aircraft_id = a.id
        LEFT JOIN pilots p2 ON fe.other_pilot_id = p2.id
        ORDER BY fe.submitted_at DESC
    ''')
    flights_list = cursor.fetchall()
    conn.close()
    
    return render_template('admin/flights.html', flights=flights_list)

@app.route('/admin/ops_emails')
def admin_ops_emails():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, email, is_active, description, created_at
        FROM ops_emails
        ORDER BY created_at DESC
    ''')
    ops_emails = cursor.fetchall()
    conn.close()
    
    return render_template('admin/ops_emails.html', ops_emails=ops_emails)

@app.route('/admin/add_ops_email', methods=['GET', 'POST'])
def admin_add_ops_email():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        description = sanitize_text_input(request.form.get('description', '').strip())
        is_active = request.form.get('is_active') == 'on'
        
        if not email:
            flash('Email address is required', 'error')
            return redirect(url_for('admin_add_ops_email'))
        
        if '@' not in email or '.' not in email:
            flash('Please enter a valid email address', 'error')
            return redirect(url_for('admin_add_ops_email'))
        
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO ops_emails (email, description, is_active)
                VALUES (?, ?, ?)
            ''', (email, description or None, is_active))
            conn.commit()
            flash(f'Ops email {email} added successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: This email address already exists!', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('admin_ops_emails'))
    
    return render_template('admin/add_ops_email.html')

@app.route('/admin/toggle_ops_email/<int:email_id>')
def admin_toggle_ops_email(email_id):
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT is_active, email FROM ops_emails WHERE id = ?", (email_id,))
    result = cursor.fetchone()
    
    if result:
        new_status = not result[0]
        cursor.execute("UPDATE ops_emails SET is_active = ? WHERE id = ?", (new_status, email_id))
        conn.commit()
        status_text = "activated" if new_status else "deactivated"
        flash(f'Ops email {result[1]} {status_text} successfully!', 'success')
    else:
        flash('Ops email not found!', 'error')
    
    conn.close()
    return redirect(url_for('admin_ops_emails'))

@app.route('/admin/delete_ops_email/<int:email_id>')
def admin_delete_ops_email(email_id):
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT email FROM ops_emails WHERE id = ?", (email_id,))
    email_result = cursor.fetchone()
    
    if email_result:
        cursor.execute("DELETE FROM ops_emails WHERE id = ?", (email_id,))
        conn.commit()
        flash(f'Ops email {email_result[0]} deleted successfully!', 'success')
    else:
        flash('Ops email not found!', 'error')
    
    conn.close()
    return redirect(url_for('admin_ops_emails'))

@app.route('/pilot_monthly', methods=['GET'])
def admin_pilot_monthly():
    """Generate monthly report for a specific pilot."""
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    pilot_id = request.args.get('pilot_id')
    month = request.args.get('month')
    
    if not pilot_id or not month:
        flash('Please select a pilot and month', 'error')
        return redirect(url_for('admin_dashboard'))
    
    try:
        year, month = map(int, month.split('-'))
    except (ValueError, AttributeError):
        flash('Invalid date format', 'error')
        return redirect(url_for('admin_dashboard'))
    
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    # Get pilot details
    cursor.execute("SELECT name FROM pilots WHERE id = ?", (pilot_id,))
    pilot = cursor.fetchone()
    
    if not pilot:
        flash('Pilot not found', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Get all pilots for the dropdown
    cursor.execute("SELECT id, name FROM pilots WHERE is_admin = FALSE ORDER BY name")
    all_pilots = cursor.fetchall()
    
    # Calculate date range
    month_start = datetime(year, month, 1).date()
    if month == 12:
        month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Get flight entries for the selected month
    cursor.execute('''
        SELECT fe.flight_date, a.registration, fe.logsheet_number, fe.q_number,
               fe.flight_type, fe.airtime, fe.amount_earned, fe.short_notice,
               fe.rate_applied, p2.name as other_pilot
        FROM flight_entries fe
        JOIN aircraft a ON fe.aircraft_id = a.id
        LEFT JOIN pilots p2 ON fe.other_pilot_id = p2.id
        WHERE fe.pilot_id = ? AND fe.flight_date BETWEEN ? AND ?
        ORDER BY fe.flight_date, a.registration
    ''', (pilot_id, month_start, month_end))
    
    entries = cursor.fetchall()
    
    # Calculate totals
    total_flights = len(entries)
    total_hours = sum(entry[5] for entry in entries)
    total_earnings = sum(entry[6] for entry in entries)
    
    conn.close()
    
    pilot_summary = {
        'name': pilot[0],
        'month': month_start.strftime('%B %Y'),
        'flights': entries,
        'total_flights': total_flights,
        'total_hours': total_hours,
        'total_earnings': total_earnings
    }
    
    return render_template('admin/dashboard.html',
                         pilot_summary=pilot_summary,
                         all_pilots=all_pilots,
                         current_month=month_start.strftime('%Y-%m'))

@app.route('/admin/generate_reports')
def admin_generate_reports():
    if 'pilot_id' not in session or not session.get('is_admin'):
        return redirect(url_for('pilot_login'))
    
    try:
        now = datetime.now()
        last_month = now.replace(day=1) - timedelta(days=1)
        
        sender = EmailSender()
        success = sender.send_monthly_reports(last_month.year, last_month.month)
        
        if success:
            flash(f'Monthly reports generated and sent for {last_month.strftime("%B %Y")}!', 'success')
        else:
            flash('Error generating reports. Check email configuration.', 'error')
    except Exception as e:
        logger.error(f"Error in admin_generate_reports: {e}")
        flash('Error generating reports.', 'error')
    
    return redirect(url_for('admin_dashboard'))

# ==============================================================================
# ERROR HANDLERS
# ==============================================================================

@app.errorhandler(404)
def not_found(error):
    if 'pilot_id' in session:
        if session.get('is_admin'):
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('pilot_dashboard'))
    return redirect(url_for('pilot_login'))

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {error}")
    flash('An internal error occurred. Please try again.', 'error')
    if 'pilot_id' in session:
        if session.get('is_admin'):
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('pilot_dashboard'))
    return redirect(url_for('pilot_login'))

@app.errorhandler(413)
def file_too_large(error):
    flash('File too large! Maximum file size is 16MB.', 'error')
    return redirect(request.url)

@app.errorhandler(429)
def rate_limit_handler(e):
    flash('Too many requests. Please try again later.', 'error')
    return redirect(url_for('pilot_login'))

# ==============================================================================
# SCHEDULED TASKS
# ==============================================================================

@scheduler.scheduled_job('cron', id='send_monthly_reports', day=1, hour=2, minute=0)
def scheduled_monthly_reports():
    """Send monthly reports on the 1st of each month at 2:00 AM"""
    logger.info("Starting scheduled monthly report generation")
    try:
        now = datetime.now()
        last_month = now.replace(day=1) - timedelta(days=1)
        
        sender = EmailSender()
        sender.send_monthly_reports(last_month.year, last_month.month)
        logger.info("Scheduled monthly reports completed successfully")
    except Exception as e:
        logger.error(f"Error in scheduled monthly reports: {e}")

# ==============================================================================
# STARTUP AND MAIN
# ==============================================================================

if __name__ == '__main__':
    # Initialize scheduler
    scheduler = BackgroundScheduler(daemon=True)
    scheduler.add_job(scheduled_monthly_reports, 'cron', day=1, hour=2, minute=0)

    def start_scheduler():
        """Start the scheduler if it's not already running."""
        if not scheduler.running:
            try:
                scheduler.start()
                logger.info("Scheduler started successfully")
            except Exception as e:
                logger.error(f"Failed to start scheduler: {e}")

    # Check security configuration first
    check_security_config()
    
    # Initialize admin user on startup
    init_admin_user()
    
    # Start the scheduler if not already running
    start_scheduler()
    
    # Configure and run the app
    port = int(os.getenv('PORT', 5000))
    debug_mode = os.getenv('FLASK_ENV') == 'development'
    
    # Force HTTP in development
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    
    logger.info(f"🚀 Starting Frosty's Pilot Portal on port {port}")
    logger.info(f"🔒 Security features: CSRF, Rate Limiting, Secure Headers")
    logger.info(f"📧 Email system: {'Configured' if os.getenv('GMAIL_USER') and os.getenv('GMAIL_PASSWORD') else 'Not configured'}")
    logger.info(f"❄️  Frosty's Operations: Professional Flight Services")
    
    try:
        # Configure SSL context
        ssl_context = None
        cert_path = os.path.join('certs', 'cert.pem')
        key_path = os.path.join('certs', 'key.pem')
        
        if os.path.exists(cert_path) and os.path.exists(key_path):
            ssl_context = (cert_path, key_path)
            logger.info("Using existing SSL certificate")
        else:
            logger.warning("No SSL certificate found. Would you like to generate a self-signed certificate? (y/n)")
            logger.warning("Note: For production, use a certificate from a trusted CA.")
            
            # Generate a self-signed certificate if not exists
            from OpenSSL import crypto
            
            # Create a key pair
            k = crypto.PKey()
            k.generate_key(crypto.TYPE_RSA, 2048)
            
            # Create a self-signed cert
            cert = crypto.X509()
            cert.get_subject().C = "US"
            cert.get_subject().ST = "State"
            cert.get_subject().L = "City"
            cert.get_subject().O = "Your Organization"
            cert.get_subject().OU = "Your Organizational Unit"
            cert.get_subject().CN = "localhost"
            cert.set_serial_number(1000)
            cert.gmtime_adj_notBefore(0)
            cert.gmtime_adj_notAfter(365*24*60*60)  # Valid for 1 year
            cert.set_issuer(cert.get_subject())
            cert.set_pubkey(k)
            cert.sign(k, 'sha256')
            
            # Save certificate
            with open(cert_path, "wb") as f:
                f.write(crypto.dump_certificate(crypto.FILETYPE_PEM, cert))
            with open(key_path, "wb") as f:
                f.write(crypto.dump_privatekey(crypto.FILETYPE_PEM, k))
                
            ssl_context = (cert_path, key_path)
            logger.info(f"Generated self-signed certificate at {cert_path}")
        
        # Run the app with HTTPS
        app.run(
            host='0.0.0.0',
            port=port,
            debug=debug_mode,
            ssl_context=ssl_context,
            use_reloader=False
        )
    except KeyboardInterrupt:
        logger.info("Shutting down...")
        scheduler.shutdown()
        raise